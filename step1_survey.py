"""
============================================================
 Step 1 — 서울 한옥 에어비앤비 전수조사
============================================================
【목적】
  서울 5개 구의 전체 행정동을 하나씩 검색하여
  '한옥' 키워드가 포함된 숙소의 listing_id 목록을 수집합니다.
  가격·예약률은 수집하지 않으며, 오직 "한옥 목록 완성"이 목표.

【출력】
  한옥_목록.xlsx  ← Step 2에서 이 파일을 읽어 상세 데이터 수집

【실행 중 복구】
  동(洞)마다 중간 저장 → 실행 도중 멈춰도 한옥_목록_진행중.csv 보존
  재실행 시 이미 조사한 동은 건너뜀 (재개 기능)

설치 (처음 한 번만):
    pip install playwright pandas openpyxl
    playwright install chromium

실행:
    python step1_survey.py
============================================================
"""

import asyncio
import base64
import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ============================================================
# ⚙️  설정
# ============================================================
SAVE_DIR          = Path(__file__).parent
PROGRESS_FILE     = SAVE_DIR / "한옥_목록_진행중.csv"   # 중간 저장
OUTPUT_FILE       = SAVE_DIR / "한옥_목록.xlsx"          # 최종 출력

MAX_PAGES_PER_DONG = 15    # 동당 최대 페이지 (× 18개 = 최대 270개)
DELAY_PAGE         = 0.5   # 페이지 간 대기(초)
DELAY_DONG         = 1.0   # 동 전환 대기(초)

# 한옥 판별 키워드
HANOK_KEYWORDS = [
    "한옥", "hanok", "韓屋",
    "기와", "고택", "전통가옥", "전통 가옥",
    "한옥스테이", "한옥 스테이",
]

# ============================================================
# 서울 5개 구 × 행정동 목록
# ============================================================
DONG_LIST = {
    "종로구": [
        "청운효자동", "사직동", "삼청동", "부암동", "평창동",
        "무악동", "교남동", "가회동", "종로1·2·3·4가동", "종로5·6가동",
        "이화동", "혜화동", "창신1동", "창신2동", "창신3동",
        "숭인1동", "숭인2동",
    ],
    "중구": [
        "소공동", "회현동", "명동", "필동", "장충동",
        "광희동", "을지로동", "신당동", "다산동", "약수동",
        "청구동", "신당5동", "동화동", "황학동", "중림동",
    ],
    "성북구": [
        "성북동", "삼선동", "동선동", "안암동", "보문동",
        "정릉1동", "정릉2동", "정릉3동", "정릉4동",
        "길음1동", "길음2동", "종암동",
        "월곡1동", "월곡2동",
        "장위1동", "장위2동", "장위3동", "석관동",
    ],
    "동대문구": [
        "용신동", "제기동", "전농1동", "전농2동",
        "답십리1동", "답십리2동", "장안1동", "장안2동",
        "청량리동", "회기동", "휘경1동", "휘경2동",
        "이문1동", "이문2동",
    ],
    "서대문구": [
        "천연동", "북아현동",
        "홍제1동", "홍제2동", "홍제3동",
        "남가좌1동", "남가좌2동", "북가좌1동", "북가좌2동",
        "홍은1동", "홍은2동",
        "연희동", "창천동", "신촌동", "충현동", "모래내로",
    ],
}

TOTAL_DONG = sum(len(v) for v in DONG_LIST.values())

# ============================================================
# 🌐  JavaScript: 검색 API 호출
# ============================================================
JS_FETCH = """
async ([apiKey, requestBody, query, cursor]) => {
    try {
        const body   = JSON.parse(requestBody);
        const vars   = body.variables || {};
        const req    = vars.staysSearchRequest || vars.request || {};
        const params = req.rawParams || [];

        const setParam = (name, value) => {
            const found = params.find(p => p.filterName === name);
            if (found) { found.filterValues = [value]; }
            else        { params.push({ filterName: name, filterValues: [value] }); }
        };
        const removeParam = (name) => {
            const idx = params.findIndex(p => p.filterName === name);
            if (idx >= 0) params.splice(idx, 1);
        };

        // 날짜 제거 (최대한 많은 숙소 노출)
        removeParam('checkin');
        removeParam('checkout');
        setParam('query', query);

        if (cursor) { setParam('cursor', cursor); }
        else        { removeParam('cursor'); }

        req.rawParams = params;

        const resp = await fetch(
            '/api/v3/StaysSearch?operationName=StaysSearch&locale=ko&currency=KRW',
            {
                method: 'POST',
                headers: {
                    'Content-Type':              'application/json',
                    'X-Airbnb-API-Key':          apiKey,
                    'X-Airbnb-GraphQL-Platform': 'web',
                    'Accept':                    'application/json',
                },
                body: JSON.stringify(body),
            }
        );
        if (!resp.ok) return { error: resp.status };
        return await resp.json();
    } catch(e) { return { error: e.toString() }; }
}
"""

# ============================================================
# 🔍  응답 파싱
# ============================================================

def parse_results(data: dict) -> tuple[list[dict], list[str]]:
    listings, cursors = [], []
    try:
        results = data["data"]["presentation"]["staysSearch"]["results"]
    except (KeyError, TypeError):
        return listings, cursors

    cursors = (results.get("paginationInfo") or {}).get("pageCursors", [])

    for result in results.get("searchResults", []):
        raw_id = (result.get("demandStayListing") or {}).get("id", "")
        listing_id = ""
        if raw_id:
            try:    listing_id = base64.b64decode(raw_id).decode().split(":")[-1]
            except: listing_id = raw_id
        if not listing_id:
            continue

        name = (
            (result.get("nameLocalized") or {}).get("localizedStringWithTranslationPreference")
            or result.get("subtitle")
            or result.get("title") or ""
        )

        # 별점 / 후기
        rating, review_cnt = "", ""
        a11y = result.get("avgRatingA11yLabel", "")
        if a11y and a11y not in ("신규 숙소", "신규"):
            m = re.search(r"(\d+\.\d+)", a11y)
            if m: rating = m.group(1)
            m = re.search(r"후기\s*(\d[\d,]*)\s*개", a11y)
            if m: review_cnt = m.group(1).replace(",", "")

        # 배지
        superhost, guest_fav = "", ""
        for badge in result.get("badges", []):
            t = badge.get("text", "")
            if "슈퍼호스트" in t or "Superhost" in t: superhost = "Y"
            if "게스트 선호" in t or "Guest favorite" in t: guest_fav = "Y"

        listings.append({
            "listing_id":  listing_id,
            "숙소_이름":   name,
            "링크":        f"https://www.airbnb.co.kr/rooms/{listing_id}",
            "별점":        float(rating) if rating else "",
            "후기_수":     int(review_cnt) if review_cnt else "",
            "슈퍼호스트":  superhost,
            "게스트_픽":   guest_fav,
        })

    return listings, cursors


def is_hanok(listing: dict) -> bool:
    text = listing.get("숙소_이름", "").lower()
    return any(kw.lower() in text for kw in HANOK_KEYWORDS)


# ============================================================
# 💾  중간 저장 / 재개
# ============================================================

def load_progress() -> tuple[dict[str, dict], set[str]]:
    """이전 실행 결과 불러오기. 반환: (hanok_dict, done_dong_set)"""
    hanok_dict = {}
    done_dong  = set()

    if PROGRESS_FILE.exists():
        try:
            df = pd.read_csv(PROGRESS_FILE, encoding="utf-8-sig")
            for _, row in df.iterrows():
                lid = str(row.get("listing_id", ""))
                if lid:
                    hanok_dict[lid] = row.to_dict()
                dong_key = f"{row.get('구','')}_{row.get('동','')}"
                done_dong.add(dong_key)
            print(f"  📂 이전 진행 불러옴: {len(hanok_dict)}개 한옥, {len(done_dong)}개 동 완료")
        except Exception as e:
            print(f"  ⚠️  진행 파일 읽기 실패: {e}")

    return hanok_dict, done_dong


def save_progress(hanok_dict: dict):
    if not hanok_dict:
        return
    pd.DataFrame(hanok_dict.values()).to_csv(
        PROGRESS_FILE, index=False, encoding="utf-8-sig"
    )


# ============================================================
# 📊  최종 엑셀 저장
# ============================================================

def save_excel(hanok_dict: dict):
    rows = list(hanok_dict.values())
    # 후기 수 내림차순 정렬
    rows.sort(key=lambda x: (-(x.get("후기_수") or 0), -(float(x.get("별점") or 0))))

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for i, r in enumerate(rows, 1):
        r["순위"]     = i
        r["수집_시각"] = now

    COLUMNS = [
        "순위", "구", "동", "숙소_이름", "링크",
        "별점", "후기_수", "슈퍼호스트", "게스트_픽", "수집_시각",
    ]
    COL_DISPLAY = {
        "순위": "순위", "구": "구(區)", "동": "동(洞)",
        "숙소_이름": "숙소 이름", "링크": "링크 (URL)",
        "별점": "별점", "후기_수": "후기 수",
        "슈퍼호스트": "슈퍼호스트", "게스트_픽": "게스트 픽",
        "수집_시각": "수집 시각",
    }
    COL_WIDTHS = {
        "순위": 6, "구(區)": 10, "동(洞)": 14,
        "숙소 이름": 40, "링크 (URL)": 10,
        "별점": 7, "후기 수": 8,
        "슈퍼호스트": 10, "게스트 픽": 10, "수집 시각": 16,
    }

    df = pd.DataFrame(rows, columns=COLUMNS).rename(columns=COL_DISPLAY)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="한옥 목록")
        ws = writer.sheets["한옥 목록"]

        hfill = PatternFill("solid", fgColor="8B4513")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        thin  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        for col_idx, col in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill, c.font = hfill, hfont
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin

        for row_idx in range(2, ws.max_row + 1):
            fill = (PatternFill("solid", fgColor="FDF5E6")
                    if row_idx % 2 == 0
                    else PatternFill("solid", fgColor="FFFFFF"))
            for col_idx in range(1, ws.max_column + 1):
                c   = ws.cell(row=row_idx, column=col_idx)
                hdr = ws.cell(row=1, column=col_idx).value
                c.fill, c.border = fill, thin
                c.alignment = Alignment(vertical="center", wrap_text=True)
                if hdr == "링크 (URL)" and c.value:
                    c.hyperlink, c.value = c.value, "🔗 링크"
                    c.font = Font(color="0563C1", underline="single")

        for col_idx, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col, 14)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # 구·동별 요약
        summary = (
            df.groupby(["구(區)", "동(洞)"])
            .agg(한옥수=("숙소 이름", "count"))
            .reset_index()
            .sort_values(["구(區)", "한옥수"], ascending=[True, False])
        )
        summary.to_excel(writer, index=False, sheet_name="구·동별 요약")

    print(f"\n✅ 최종 저장 완료: {OUTPUT_FILE.name}  ({len(rows)}개 한옥)")


# ============================================================
# 🚀  메인
# ============================================================

async def main():
    print("=" * 62)
    print(" 🏯 Step 1 — 서울 한옥 에어비앤비 전수조사")
    print(f"    대상: {len(DONG_LIST)}개 구 / {TOTAL_DONG}개 행정동")
    print("=" * 62)

    # 이전 진행 불러오기
    all_hanok, done_dong = load_progress()

    captured = {"api_key": None, "request_body": None}

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
        )
        ctx = await browser.new_context(
            viewport={"width": 1440, "height": 900},
            locale="ko-KR", timezone_id="Asia/Seoul",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        await ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
            "window.chrome={runtime:{}};"
        )
        page = await ctx.new_page()

        async def on_req(request):
            if "StaysSearch" in request.url and not captured["api_key"]:
                key = request.headers.get("x-airbnb-api-key", "")
                if key:
                    captured["api_key"] = key
                    try: captured["request_body"] = request.post_data
                    except: pass

        page.on("request", on_req)

        print("\n[1단계] 에어비앤비 접속 중...")
        await page.goto(
            "https://www.airbnb.co.kr/s/서울/homes?tab_id=home_tab&refinement_paths%5B%5D=%2Fhomes",
            wait_until="load", timeout=60000,
        )
        await asyncio.sleep(2)

        for _ in range(30):
            if captured["api_key"] and captured["request_body"]: break
            await asyncio.sleep(0.5)

        if not captured["api_key"]:
            print("❌ API 키 캡처 실패.")
            await browser.close()
            return
        print("        ✅ API 키 캡처 성공!\n")

        # ── 동 순회 ──────────────────────────────────────────
        print(f"[2단계] {TOTAL_DONG}개 동 순회 시작\n")

        dong_seq  = 0
        for gu, dongs in DONG_LIST.items():
            for dong in dongs:
                dong_seq += 1
                dong_key  = f"{gu}_{dong}"
                query     = f"서울 {gu} {dong}"

                # 이미 완료한 동이면 건너뜀
                if dong_key in done_dong:
                    print(f"  [{dong_seq:03d}/{TOTAL_DONG}] {gu} {dong}: 건너뜀 (이미 완료)")
                    continue

                dong_total = 0
                dong_hanok = 0

                result = await page.evaluate(
                    JS_FETCH,
                    [captured["api_key"], captured["request_body"], query, None]
                )
                if not result or result.get("error"):
                    print(f"  [{dong_seq:03d}/{TOTAL_DONG}] {gu} {dong}: 오류 — 건너뜀")
                    continue

                items, cursors = parse_results(result)
                dong_total += len(items)
                for item in items:
                    if is_hanok(item) and item["listing_id"] not in all_hanok:
                        item["구"] = gu
                        item["동"] = dong
                        all_hanok[item["listing_id"]] = item
                        dong_hanok += 1

                for p_idx in range(1, min(MAX_PAGES_PER_DONG, len(cursors))):
                    r2 = await page.evaluate(
                        JS_FETCH,
                        [captured["api_key"], captured["request_body"],
                         query, cursors[p_idx]]
                    )
                    if not r2 or r2.get("error"): break
                    items2, _ = parse_results(r2)
                    if not items2: break
                    dong_total += len(items2)
                    for item in items2:
                        if is_hanok(item) and item["listing_id"] not in all_hanok:
                            item["구"] = gu
                            item["동"] = dong
                            all_hanok[item["listing_id"]] = item
                            dong_hanok += 1
                    await asyncio.sleep(DELAY_PAGE)

                done_dong.add(dong_key)
                save_progress(all_hanok)   # 동마다 저장

                marker = "🏯" if dong_hanok > 0 else "  "
                print(f"  {marker}[{dong_seq:03d}/{TOTAL_DONG}] {gu} {dong}: "
                      f"조회 {dong_total}개 → 한옥 {dong_hanok}개 (누적 {len(all_hanok)}개)")

                await asyncio.sleep(DELAY_DONG)

        await browser.close()

    # ── 최종 저장 ─────────────────────────────────────────────
    if not all_hanok:
        print("\n❌ 한옥 숙소를 찾지 못했습니다.")
        return

    print(f"\n[3단계] 최종 엑셀 저장 중...")
    save_excel(all_hanok)

    # 진행 파일 삭제 (완료)
    if PROGRESS_FILE.exists():
        PROGRESS_FILE.unlink()
        print(f"  🗑️  진행 파일 삭제 완료")

    # 요약
    print("\n" + "=" * 62)
    print(" 📊 전수조사 완료")
    print("=" * 62)
    print(f"  총 한옥 숙소: {len(all_hanok)}개")
    print(f"  조사 완료 동: {len(done_dong)}개 / {TOTAL_DONG}개")

    by_gu = {}
    for item in all_hanok.values():
        by_gu[item.get("구", "?")] = by_gu.get(item.get("구", "?"), 0) + 1
    print("\n  구별 결과:")
    for gu, cnt in sorted(by_gu.items(), key=lambda x: -x[1]):
        bar = "■" * cnt
        print(f"    {gu:6s}  {cnt:3d}개  {bar}")

    print(f"\n  → {OUTPUT_FILE.name}")
    print("  ※ Step 2에서 이 파일을 읽어 상세 데이터를 수집합니다.")


if __name__ == "__main__":
    asyncio.run(main())
