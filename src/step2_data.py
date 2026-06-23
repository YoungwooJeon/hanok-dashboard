"""
============================================================
 Step 2 — 한옥 스테이 상세 데이터 수집  (v3 — 실제 API 구조 반영판)
============================================================
【전제】
  Step 1이 완료되어 '한옥_목록.xlsx'가 존재해야 합니다.

【배경 — 왜 다시 작성했는가】
  에어비앤비는 v2 API(/api/v2/listings, /api/v2/calendar_months,
  /api/v2/reviews)를 완전히 폐지했습니다. 전부 404가 떨어집니다.
  실제로 살아있는 API는 아래 3개의 v3 GraphQL 쿼리입니다.

    · StaysPdpSections      — 숙소 기본 정보(호스트, 구성, 정책 등)
                              ⚠ 스크롤 위치마다 "여러 번" 호출되고,
                                매번 다른 섹션이 실려 옵니다.
                                → 모두 누적해서 합쳐야 함.
    · StaysPdpReviewsQuery  — 후기 (comments, createdAt)
    · PdpAvailabilityCalendar — 캘린더(예약률 계산용)

【검증 상태 — 투명하게 공개】
  실제 debug JSON 파일을 분석해 다음은 "확인됨":
    - 호스트 이름        ("호스트: RimJae 님" 텍스트에서 파싱)
    - 호스팅 경력/시작    ("호스팅 경력 9년" 텍스트에서 파싱)
    - 최대 인원, 구성     ("최대 인원 6명", "침실 3개" 등 텍스트 파싱)
    - 슈퍼호스트
    - 후기 텍스트/날짜    (StaysPdpReviewsQuery)

  다음은 "미확인 — 최선 노력(best-effort) 파서로 처리":
    - 최소 박수, 청소비, 추가인원 요금, 소개글
      → 지금까지 받은 모든 debug 응답에 해당 필드가 보이지 않았습니다.
        키 이름 패턴으로 전체 JSON을 재귀 탐색하지만, 못 찾으면
        빈 값으로 남고 콘솔에 "미확인" 표시가 뜹니다.

  1박/주중/주말 가격 (2026-06-18 재작업):
    → 처음엔 검색(StaysSearch) API에 listingId를 끼워 넣어 가져오려
      했으나, 그 API가 필터를 무시하고 전세계 무관한 숙소를 섞어서
      반환한다는 걸 debug 응답으로 확인 → 폐기.
    → 지금은 숙소 상세 페이지(PDP)를 직접 열어 그 페이지가 응답한
      BookItSection의 실제 가격을 읽는다. 요청한 날짜에 예약이 막혀
      있으면(품절/차단) 그 페이지엔 가격이 없으므로, 캘린더에서 실제
      예약 가능한 가장 가까운 날짜를 찾아 한 번 더 시도한다.
      그래도 못 찾으면(연중 예약 불가 등) 빈 값으로 둔다.
    - 캘린더(예약률 30일 / 가격 fallback용 가용일 탐색)
      → PdpAvailabilityCalendar 응답에서 calendarDate/available 필드를
        확인했다. 단, 이 API의 price 필드는 항상 null로 와서(호스트
        스마트프라이싱 설정과 무관하게) 가격 출처로는 쓸 수 없었다 —
        그래서 가격은 PDP 직접 조회로 얻는다.

  실행 후 1번째 숙소의 원본 응답을 debug2_*.json 으로 저장하니,
  결과에서 빈 칸이 많이 보이면 그 파일을 보고 다시 알려주세요.
  → 정확한 키 경로를 찾아 파서를 보강하겠습니다.

설치:
    pip install playwright pandas openpyxl
    playwright install chromium

실행:
    python step2_data.py
============================================================
"""

import asyncio
import json
import re
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ============================================================
# ⚙️  설정
# ============================================================
SAVE_DIR     = Path(__file__).parent
INPUT_FILE   = SAVE_DIR / "한옥_목록.xlsx"
OUTPUT_FILE  = SAVE_DIR / "한옥_분석.xlsx"

DELAY_DETAIL   = 0.8
SCROLL_STEPS   = 10     # lazy-load 섹션을 모두 유발하기 위한 스크롤 횟수
SCROLL_WAIT    = 0.35   # 스크롤 간 대기(초)

# 가격 기준 날짜 (모두 "1박" 기준 — 박당 가격을 비교하기 위해 통일)
_TODAY   = datetime.now()
NIGHTS   = 1
CHECKIN  = (_TODAY + timedelta(days=1)).strftime("%Y-%m-%d")
CHECKOUT = (_TODAY + timedelta(days=1 + NIGHTS)).strftime("%Y-%m-%d")

def _next_weekday(dow: int) -> datetime:
    days = (dow - _TODAY.weekday()) % 7 or 7
    return _TODAY + timedelta(days=days)

_TUE        = _next_weekday(1)
WEEKDAY_IN  = _TUE.strftime("%Y-%m-%d")
WEEKDAY_OUT = (_TUE + timedelta(days=1)).strftime("%Y-%m-%d")

_SAT        = _next_weekday(5)
WEEKEND_IN  = _SAT.strftime("%Y-%m-%d")
WEEKEND_OUT = (_SAT + timedelta(days=1)).strftime("%Y-%m-%d")

# ============================================================
# 🔑  후기 키워드 그룹
# ============================================================
KEYWORD_GROUPS = {
    "마당·정원":     ["마당", "정원", "뜰", "마당에서", "마당이"],
    "툇마루·평상":   ["툇마루", "평상", "대청마루", "마루", "처마"],
    "다락방":        ["다락", "다락방", "로프트"],
    "루프탑·옥상":   ["루프탑", "옥상", "옥상에서", "옥상 뷰"],
    "독채·프라이빗": ["독채", "단독", "프라이빗", "전세", "우리만", "혼자"],
    "아궁이·장작":   ["아궁이", "장작", "화로", "모닥불", "불멍"],
    "온돌·구들":     ["온돌", "구들", "바닥 난방", "바닥이 따뜻", "바닥 온기"],
    "전통 가구·소품":["전통 가구", "도자기", "항아리", "병풍", "고가구", "민화", "전통 소품"],
    "기와·한옥 외관":["기와", "처마", "한옥 느낌", "한옥 외관", "전통 건축"],
    "한복 체험":     ["한복", "한복 대여", "한복 입고"],
    "다도·차":       ["다도", "차 한잔", "전통차", "다실", "티 세트"],
    "바베큐·화로":   ["바베큐", "bbq", "그릴", "화로구이", "고기 구워"],
    "요리·주방 체험":["직접 요리", "주방 이용", "밥 해먹", "요리해"],
    "조식 제공":     ["조식", "아침 식사", "아침을 제공", "아침 차려", "조반"],
    "웰컴 푸드":     ["웰컴", "간식", "과일", "음료 제공", "다과", "떡", "전통 과자"],
    "막걸리·전통주": ["막걸리", "전통주", "동동주", "약주", "술 제공"],
    "족욕·반신욕":   ["족욕", "반신욕", "족욕기", "발 담그"],
    "노천탕·욕조":   ["노천탕", "야외 욕조", "자쿠지", "스파", "히노키"],
    "사우나·찜질":   ["사우나", "찜질", "한증"],
    "수영장":        ["수영장", "풀", "pool", "수영"],
    "경복궁·고궁 뷰":["경복궁", "창덕궁", "창경궁", "고궁", "궁궐"],
    "산·자연 뷰":    ["북한산", "인왕산", "산 뷰", "산이 보", "자연 뷰", "숲"],
    "한양도성·성곽": ["성곽", "한양도성", "도성", "성벽"],
    "야경·도심 뷰":  ["야경", "도심 뷰", "서울 뷰", "야경이", "불빛"],
    "별·달 감상":    ["별 보", "별이 보", "달 보", "달빛", "밤하늘"],
    "꽃·단풍·눈":    ["벚꽃", "단풍", "눈 오", "눈 내", "꽃이 피"],
    "자연 소리":     ["새소리", "물소리", "바람 소리", "자연 소리"],
    "커플·로맨틱":   ["커플", "로맨틱", "데이트", "둘이서", "프로포즈", "기념일"],
    "가족·키즈":     ["가족", "아이", "어린이", "키즈", "아이들", "아기"],
    "반려동물":      ["강아지", "고양이", "반려동물", "펫", "pet", "댕댕"],
    "혼자 여행":     ["혼자", "혼행", "1인", "나 혼자", "혼자서"],
    "픽업·셔틀":     ["픽업", "셔틀", "데리러", "공항", "터미널"],
    "자전거·킥보드": ["자전거", "킥보드", "자전거 대여"],
    "주차":          ["주차", "주차장", "주차 가능", "차 세워"],
}

AMENITY_CHECKS = {
    "에어컨":  ["에어컨", "에어 컨", "air conditioning", "air_conditioning"],
    "주방":    ["주방", "부엌", "kitchen", "kitchenette"],
    "세탁기":  ["세탁기", "washer", "washing machine"],
    "주차":    ["주차", "parking", "free parking"],
    "와이파이": ["wifi", "wi-fi", "와이파이", "무선 인터넷"],
    "TV":      ["tv", "television", "티비", "텔레비전"],
    "욕조":    ["욕조", "bathtub", "bath tub", "soaking tub"],
    "난방":    ["난방", "heating", "온돌", "보일러"],
}

# ============================================================
# 🔍  범용 재귀 탐색 헬퍼
#     (정확한 JSON 경로를 모르는 필드를 찾기 위한 안전망)
# ============================================================

def _walk_dicts(obj):
    """중첩된 구조 안의 모든 dict를 재귀적으로 순회"""
    if isinstance(obj, dict):
        yield obj
        for v in obj.values():
            yield from _walk_dicts(v)
    elif isinstance(obj, list):
        for item in obj:
            yield from _walk_dicts(item)


def find_by_typename(obj, typename):
    """__typename이 일치하는 모든 dict 반환"""
    return [d for d in _walk_dicts(obj) if d.get("__typename") == typename]


def find_keys_containing(obj, substr_lower):
    """키 이름에 substr(소문자)이 포함된 모든 (key, value) 반환"""
    out = []
    for d in _walk_dicts(obj):
        for k, v in d.items():
            if substr_lower in k.lower():
                out.append((k, v))
    return out


# ============================================================
# 🔍  파서 — 가격
#
# ⚠ 버그 수정 내역 (2026-06-18):
#   1차 시도: StaysSearch(검색) API에 listingId 필터를 끼워 넣어 특정
#   숙소의 가격만 받아오려 했으나, 실제 debug 응답을 까보니 그 API는
#   listingId 필터를 사실상 무시하고 전세계 다른 나라 숙소들이 뒤섞인
#   "일반 인기 검색 결과"를 그대로 돌려준다는 것을 확인했다(브라질·베트남
#   숙소 등이 한국 한옥 검색에 섞여 나옴). 즉 이 API로는 특정 숙소 1곳의
#   가격을 절대 신뢰성 있게 가져올 수 없다 — 키 경로를 고쳐도 해결 불가능한
#   구조적 한계였다.
#   → 최종 방식: 숙소 상세 페이지(PDP)를 ?check_in=...&check_out=...로
#     직접 열어서, 그 페이지가 실제로 응답한 BookItSection 안의
#     structuredDisplayPrice를 읽는다. 단, 그 날짜에 예약이 불가능한
#     숙소는 가격이 비어 있으므로(available=false), 캘린더에서 실제로
#     예약 가능한 가장 가까운 날짜를 찾아 한 번 더 시도한다.
# ============================================================

def parse_price_from_sections(sections_list: list, nights: int = 1):
    """StaysPdpSections 응답들 안의 BookItSection에서 '예약 가능한' 가격을 추출.
    available=false(해당 날짜 예약 불가)인 섹션은 가격이 없으므로 건너뛴다."""
    for sections in sections_list:
        for sec in find_by_typename(sections, "BookItSection"):
            if not sec.get("available"):
                continue
            sdp = sec.get("structuredDisplayPrice")
            if not sdp:
                continue
            line = sdp.get("primaryLine") or {}
            price_text = line.get("price") or line.get("discountedPrice") or ""
            m = re.search(r"[\d,]+", price_text)
            if m:
                total = int(m.group(0).replace(",", ""))
                return total // nights if nights else total
    return ""


def _find_available_window(calendar_json, today, nights: int = 1,
                            weekday_filter=None, search_days: int = 90):
    """캘린더에서 실제로 예약 가능한 연속 nights일을 찾아 (체크인, 체크아웃) 반환.
    weekday_filter: 체크인 요일을 제한할 때 사용 (0=월 ... 6=일). 못 찾으면 (None, None)."""
    if not calendar_json:
        return None, None

    days = {}
    for d in _walk_dicts(calendar_json):
        ds = d.get("calendarDate")
        if isinstance(ds, str) and _DATE_RE.match(ds):
            days[ds] = bool(d.get("available"))

    for offset in range(1, search_days):
        d0 = today + timedelta(days=offset)
        if weekday_filter is not None and d0.weekday() not in weekday_filter:
            continue
        ok = True
        for n in range(nights):
            ds = (d0 + timedelta(days=n)).strftime("%Y-%m-%d")
            if not days.get(ds, False):
                ok = False
                break
        if ok:
            return d0.strftime("%Y-%m-%d"), (d0 + timedelta(days=nights)).strftime("%Y-%m-%d")
    return None, None


async def fetch_price_for_dates(page, lid, checkin, checkout, nights: int = 1):
    """가벼운 PDP 방문으로 특정 날짜의 1박 가격을 가져온다 (스크롤 없음)."""
    cap_sections = []

    async def on_resp(response):
        if "StaysPdpSections" in response.url:
            try:
                cap_sections.append(await response.json())
            except Exception:
                pass

    page.on("response", on_resp)
    try:
        await page.goto(
            f"https://www.airbnb.co.kr/rooms/{lid}"
            f"?check_in={checkin}&check_out={checkout}&adults=1",
            wait_until="domcontentloaded", timeout=20000,
        )
        await asyncio.sleep(1.3)
    except Exception:
        pass
    finally:
        page.remove_listener("response", on_resp)

    return parse_price_from_sections(cap_sections, nights), cap_sections


# ============================================================
# 🔍  파서 — StaysPdpSections (확인됨: 호스트/구성/최대인원)
#     ⚠ 스크롤 중 여러 번 호출되므로 리스트(sections_list)를
#       모두 합쳐서 탐색해야 한다.
# ============================================================

def parse_host_and_layout(sections_list: list) -> dict:
    result = {
        "호스트_이름": "", "호스팅_시작": "",
        "최대_인원": "", "구성": "", "슈퍼호스트_확인": "",
    }
    years_ago = None
    parts = {"침실": "", "침대": "", "욕실": ""}

    for sections in sections_list:
        # 호스트 정보
        for host_sec in find_by_typename(sections, "PdpHostOverviewDefaultSection"):
            title = host_sec.get("title", "") or ""
            m = re.search(r"호스트[:：]\s*(.+?)\s*님", title)
            if m and not result["호스트_이름"]:
                result["호스트_이름"] = m.group(1).strip()
            for item in (host_sec.get("overviewItems") or []):
                t = item.get("title", "") or ""
                if "슈퍼호스트" in t:
                    result["슈퍼호스트_확인"] = "Y"
                m2 = re.search(r"호스팅\s*경력\s*(\d+)\s*년", t)
                if m2:
                    years_ago = int(m2.group(1))

        # 최대인원/구성
        for ov_sec in find_by_typename(sections, "PdpOverviewV2Section"):
            for item in (ov_sec.get("overviewItems") or []):
                t = item.get("title", "") or ""
                m3 = re.search(r"최대\s*인원\s*(\d+)\s*명", t)
                if m3 and not result["최대_인원"]:
                    result["최대_인원"] = m3.group(1)
                m4 = re.search(r"(침실|침대|욕실)\s*(\d+)\s*개", t)
                if m4:
                    parts[m4.group(1)] = m4.group(2)

    comp = " / ".join(f"{k} {v}개" for k, v in parts.items() if v)
    result["구성"] = comp
    if years_ago is not None:
        result["호스팅_시작"] = f"{datetime.now().year - years_ago}년경"
    return result


def parse_amenities(sections_list: list) -> dict:
    """어메니티 — typename에 'Amenit' 포함된 섹션을 탐색 (best-effort)"""
    texts = []
    for sections in sections_list:
        for d in _walk_dicts(sections):
            tn = d.get("__typename", "")
            if isinstance(tn, str) and "Amenit" in tn:
                t = d.get("title")
                if isinstance(t, str):
                    texts.append(t)
        for amen_sec in find_by_typename(sections, "PdpAmenitiesSection"):
            for item in (amen_sec.get("previewAmenities")
                         or amen_sec.get("amenities") or []):
                if isinstance(item, dict):
                    t = item.get("title")
                    if isinstance(t, str):
                        texts.append(t)
    blob = " ".join(texts).lower()
    result = {}
    for col, keywords in AMENITY_CHECKS.items():
        result[col] = "Y" if any(kw.lower() in blob for kw in keywords) else ""
    return result


def parse_misc_bestguess(sections_list: list) -> dict:
    """
    최소 박수 / 청소비 / 추가인원 요금 / 소개글
    ⚠ 미확인 필드 — 키 이름 패턴으로 전체 트리를 탐색.
       못 찾으면 빈 값 그대로 둔다 (거짓 데이터를 만들지 않음).
    """
    result = {"최소_박수": "", "청소비": "", "추가인원_요금": "", "소개글": ""}

    for sections in sections_list:
        if not result["최소_박수"]:
            for k, v in find_keys_containing(sections, "minnight"):
                if isinstance(v, (int, str)) and str(v).strip():
                    result["최소_박수"] = v
                    break

        if not result["청소비"]:
            for k, v in find_keys_containing(sections, "cleaningfee"):
                if v not in (None, "", 0):
                    result["청소비"] = v
                    break

        if not result["추가인원_요금"]:
            candidates = (
                find_keys_containing(sections, "extraguestfee")
                + find_keys_containing(sections, "extrapersonfee")
                + find_keys_containing(sections, "additionalguestfee")
            )
            for k, v in candidates:
                if v not in (None, "", 0):
                    result["추가인원_요금"] = v
                    break

        if not result["소개글"]:
            for desc_sec in (
                find_by_typename(sections, "PdpDescriptionSection")
                + find_by_typename(sections, "PdpDescriptionDefaultSection")
            ):
                txt = desc_sec.get("htmlDescription") or desc_sec.get("description")
                if isinstance(txt, dict):
                    txt = txt.get("htmlText", "")
                if isinstance(txt, str) and txt.strip():
                    result["소개글"] = txt.strip()[:200]
                    break
            if not result["소개글"]:
                # 키 이름 기반 보조 탐색
                for k, v in find_keys_containing(sections, "description"):
                    if isinstance(v, str) and len(v.strip()) > 20:
                        result["소개글"] = v.strip()[:200]
                        break

    return result


# ============================================================
# 🔍  파서 — PdpAvailabilityCalendar (미확인 구조 — 범용 탐색)
# ============================================================

_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def parse_booking_rate(calendar_json, today) -> str:
    if not calendar_json:
        return ""

    days = {}
    for d in _walk_dicts(calendar_json):
        date_str = d.get("calendarDate") or d.get("date")
        if isinstance(date_str, str) and _DATE_RE.match(date_str):
            avail = d.get("available")
            if avail is None:
                avail = d.get("availableForCheckin")
            if avail is None:
                avail = d.get("bookable")
            if isinstance(avail, bool):
                days[date_str] = avail

    if not days:
        return ""

    start = today + timedelta(days=1)
    end   = today + timedelta(days=31)
    window = []
    for ds, avail in days.items():
        try:
            dt = datetime.strptime(ds, "%Y-%m-%d").date()
        except Exception:
            continue
        if start <= dt < end:
            window.append(avail)

    if len(window) < 15:
        return ""

    booked = sum(1 for a in window if not a)
    return f"{round(booked / len(window) * 100)}%"


# ============================================================
# 🔍  파서 — StaysPdpReviewsQuery (확인됨)
# ============================================================

def _find_review_dicts(reviews_json):
    if not reviews_json:
        return []
    return [
        d for d in _walk_dicts(reviews_json)
        if isinstance(d.get("comments"), str) and d["comments"].strip()
    ]


def parse_reviews(reviews_json) -> str:
    revs = _find_review_dicts(reviews_json)
    if not revs:
        return ""
    all_text = " ".join(r.get("comments", "") for r in revs).lower()

    counts = {}
    for category, keywords in KEYWORD_GROUPS.items():
        cnt = sum(all_text.count(kw) for kw in keywords)
        if cnt > 0:
            counts[category] = cnt
    if not counts:
        return ""

    top5 = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:5]
    return " / ".join(f"{cat}({n})" for cat, n in top5)


def parse_reviews_oldest(reviews_json) -> str:
    revs = _find_review_dicts(reviews_json)
    dates = []
    for r in revs:
        raw = r.get("createdAt") or r.get("created_at") or ""
        try:
            dates.append(datetime.strptime(raw[:10], "%Y-%m-%d"))
        except Exception:
            pass
    if not dates:
        return ""
    return min(dates).strftime("%Y-%m")


# ============================================================
# 🚀  메인
# ============================================================

async def main():
    print("=" * 62)
    print(" 📊 Step 2 — 한옥 스테이 상세 데이터 수집 (v3)")
    print(f"    입력 파일: {INPUT_FILE.name}")
    print(f"    가격 기준: {CHECKIN} ~ {CHECKOUT} ({NIGHTS}박)")
    print("=" * 62)

    if not INPUT_FILE.exists():
        print(f"\n❌ {INPUT_FILE.name} 파일이 없습니다.")
        print("   Step 1을 먼저 완료하세요.")
        return

    from openpyxl import load_workbook as _load_wb
    _wb  = _load_wb(INPUT_FILE)
    _ws  = _wb["한옥 목록"]
    _hdr = [c.value for c in _ws[1]]

    def _col(name): return _hdr.index(name)

    listings = []
    for row in _ws.iter_rows(min_row=2, values_only=False):
        if not any(c.value for c in row):
            continue
        def v(name):
            return row[_col(name)].value if name in _hdr else ""
        link_cell = row[_col("링크 (URL)")]
        url = (link_cell.hyperlink.target
               if link_cell.hyperlink else str(link_cell.value or ""))
        m = re.search(r"/rooms/(\d+)", url)
        if not m:
            continue
        listings.append({
            "listing_id":  m.group(1),
            "구":          v("구(區)"),
            "동":          v("동(洞)"),
            "숙소_이름":   v("숙소 이름"),
            "링크":        url,
            "별점":        v("별점"),
            "후기_수":     v("후기 수"),
            "슈퍼호스트":  v("슈퍼호스트"),
            "게스트_픽":   v("게스트 픽"),
        })
    _wb.close()
    total = len(listings)
    print(f"\n    총 {total}개 한옥 숙소 로드 완료\n")
    print(f"    ⏱  예상 소요 시간: 숙소당 약 6~10초 (스크롤 포함)\n")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
        )
        ctx = await browser.new_context(
            viewport={"width": 1440, "height": 900},
            locale="ko-KR", timezone_id="Asia/Seoul",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        await ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
            "window.chrome={runtime:{}};"
        )
        # page_price: 가격 전용 가벼운 페이지 (스크롤 없이 날짜별 가격만 조회)
        # page_detail: 호스트/구성/후기/캘린더 등 전체 콘텐츠 수집용 (스크롤 포함)
        page_price  = await ctx.new_page()
        page_detail = await ctx.new_page()
        print("        ✅ 듀얼 페이지 준비 완료\n")

        print(f"[2단계] {total}개 숙소 상세 데이터 수집 중...\n")
        today   = _TODAY.date()
        results = []
        missing_counts = {"캘린더": 0, "후기": 0, "최소박수": 0, "청소비": 0,
                           "추가인원요금": 0, "소개글": 0, "호스트이름": 0}

        for idx, row in enumerate(listings, 1):
            lid  = str(row.get("listing_id", ""))
            name = str(row.get("숙소_이름", ""))[:25]

            record = {
                "listing_id":  lid,
                "구":          row.get("구", ""),
                "동":          row.get("동", ""),
                "숙소_이름":   row.get("숙소_이름", ""),
                "링크":        row.get("링크", ""),
                "별점":        row.get("별점", ""),
                "후기_수":     row.get("후기_수", ""),
                "슈퍼호스트":  row.get("슈퍼호스트", ""),
                "게스트_픽":   row.get("게스트_픽", ""),
                "호스트_이름":   "",
                "호스트_응답률": "",
                "최소_박수":     "",
                "최대_인원":     "",
                "기준_인원":     "",
                "구성":          "",
                "청소비":        "",
                "추가인원_요금": "",
                "1박_가격":      "",
                "주중_가격":     "",
                "주말_가격":     "",
                "예약률_30일":   "",
                "호스팅_시작":   "",
                "소개글":        "",
                "에어컨": "", "주방": "", "세탁기": "",
                "주차": "", "와이파이": "", "TV": "",
                "욕조": "", "난방": "",
                "후기_키워드":   "",
            }

            print(f"  [{idx:03d}/{total}] {name}")

            # ① 숙소 페이지 이동 — page_detail (호스트/구성/후기/캘린더 +
            #    CHECKIN~CHECKOUT 날짜의 BookItSection도 함께 캡처됨)
            cap = {"sections_list": [], "reviews": None, "calendar": None}

            async def on_resp(response):
                url = response.url
                try:
                    if "StaysPdpSections" in url:
                        data = await response.json()
                        cap["sections_list"].append(data)
                    elif "StaysPdpReviewsQuery" in url and cap["reviews"] is None:
                        cap["reviews"] = await response.json()
                    elif "PdpAvailabilityCalendar" in url and cap["calendar"] is None:
                        cap["calendar"] = await response.json()
                except Exception:
                    pass

            page_detail.on("response", on_resp)
            try:
                await page_detail.goto(
                    f"https://www.airbnb.co.kr/rooms/{lid}"
                    f"?check_in={CHECKIN}&check_out={CHECKOUT}&adults=1",
                    wait_until="load", timeout=30000,
                )
                await asyncio.sleep(1.2)
                # lazy-load 섹션(소개글/어메니티/캘린더 등)을 모두 유발
                for _ in range(SCROLL_STEPS):
                    await page_detail.mouse.wheel(0, 700)
                    await asyncio.sleep(SCROLL_WAIT)
                await asyncio.sleep(1.0)
            except Exception as e:
                print(f"         페이지 로딩 오류: {e}")
            finally:
                page_detail.remove_listener("response", on_resp)

            # 디버그: 첫 번째 숙소 응답 구조 저장 (검증용)
            if idx == 1:
                with open(SAVE_DIR / "debug2_sections_all.json", "w", encoding="utf-8") as f:
                    json.dump(cap["sections_list"], f, ensure_ascii=False, indent=2)
                print(f"         🔍 sections({len(cap['sections_list'])}개) 저장: debug2_sections_all.json")
                if cap["reviews"]:
                    with open(SAVE_DIR / "debug2_reviews.json", "w", encoding="utf-8") as f:
                        json.dump(cap["reviews"], f, ensure_ascii=False, indent=2)
                    print(f"         🔍 reviews 저장: debug2_reviews.json")
                if cap["calendar"]:
                    with open(SAVE_DIR / "debug2_calendar.json", "w", encoding="utf-8") as f:
                        json.dump(cap["calendar"], f, ensure_ascii=False, indent=2)
                    print(f"         🔍 calendar 저장: debug2_calendar.json")
                else:
                    print(f"         ⚠️  calendar 응답을 받지 못함 (PdpAvailabilityCalendar 미수신)")

            # ② 가격 3종 — 기본 날짜로 시도 → 예약 불가능하면 캘린더에서
            #    실제로 예약 가능한 날짜를 찾아 page_price로 한 번 더 시도
            pr1 = parse_price_from_sections(cap["sections_list"], NIGHTS)
            if not isinstance(pr1, int):
                ci, co = _find_available_window(cap["calendar"], today, nights=NIGHTS)
                if ci:
                    pr1, _ = await fetch_price_for_dates(page_price, lid, ci, co, NIGHTS)
            if isinstance(pr1, int):
                record["1박_가격"] = pr1
            else:
                missing_counts.setdefault("가격_매칭실패", 0)
                missing_counts["가격_매칭실패"] += 1

            pr2, sec2 = await fetch_price_for_dates(page_price, lid, WEEKDAY_IN, WEEKDAY_OUT, NIGHTS)
            if idx == 1:
                with open(SAVE_DIR / "debug2_price_weekday.json", "w", encoding="utf-8") as f:
                    json.dump(sec2, f, ensure_ascii=False, indent=2)
            if not isinstance(pr2, int):
                ci, co = _find_available_window(cap["calendar"], today, nights=NIGHTS,
                                                 weekday_filter={0, 1, 2, 3})
                if ci:
                    pr2, _ = await fetch_price_for_dates(page_price, lid, ci, co, NIGHTS)
            if isinstance(pr2, int):
                record["주중_가격"] = pr2

            pr3, sec3 = await fetch_price_for_dates(page_price, lid, WEEKEND_IN, WEEKEND_OUT, NIGHTS)
            if idx == 1:
                with open(SAVE_DIR / "debug2_price_weekend.json", "w", encoding="utf-8") as f:
                    json.dump(sec3, f, ensure_ascii=False, indent=2)
            if not isinstance(pr3, int):
                ci, co = _find_available_window(cap["calendar"], today, nights=NIGHTS,
                                                 weekday_filter={4, 5})
                if ci:
                    pr3, _ = await fetch_price_for_dates(page_price, lid, ci, co, NIGHTS)
            if isinstance(pr3, int):
                record["주말_가격"] = pr3

            # ③ 호스트/구성/최대인원 파싱 (확인됨)
            host_info = parse_host_and_layout(cap["sections_list"])
            record["호스트_이름"] = host_info["호스트_이름"]
            record["최대_인원"]   = host_info["최대_인원"]
            record["구성"]        = host_info["구성"]
            if host_info["호스팅_시작"]:
                record["호스팅_시작"] = host_info["호스팅_시작"]

            # ④ 어메니티 파싱 (best-effort)
            record.update(parse_amenities(cap["sections_list"]))

            # ⑤ 미확인 필드 best-effort 파싱
            misc = parse_misc_bestguess(cap["sections_list"])
            record["최소_박수"]     = misc["최소_박수"]
            record["청소비"]        = misc["청소비"]
            record["추가인원_요금"] = misc["추가인원_요금"]
            record["소개글"]        = misc["소개글"]

            # ⑥ 캘린더(예약률) 파싱
            record["예약률_30일"] = parse_booking_rate(cap["calendar"], today)

            # ⑦ 후기 파싱
            if cap["reviews"]:
                record["후기_키워드"] = parse_reviews(cap["reviews"])
                oldest = parse_reviews_oldest(cap["reviews"])
                if oldest and not record["호스팅_시작"]:
                    record["호스팅_시작"] = oldest
            else:
                missing_counts["후기"] += 1

            if not record["예약률_30일"]:   missing_counts["캘린더"] += 1
            if not record["최소_박수"]:     missing_counts["최소박수"] += 1
            if not record["청소비"]:        missing_counts["청소비"] += 1
            if not record["추가인원_요금"]: missing_counts["추가인원요금"] += 1
            if not record["소개글"]:        missing_counts["소개글"] += 1
            if not record["호스트_이름"]:   missing_counts["호스트이름"] += 1

            p_str  = f"₩{record['1박_가격']:,}"  if isinstance(record["1박_가격"],  int) else "-"
            wd_str = f"₩{record['주중_가격']:,}" if isinstance(record["주중_가격"], int) else "-"
            we_str = f"₩{record['주말_가격']:,}" if isinstance(record["주말_가격"], int) else "-"
            print(f"         호스트: {record['호스트_이름'] or '-'}  가격: {p_str}  주중: {wd_str}  주말: {we_str}"
                  f"  예약률: {record['예약률_30일'] or '-'}"
                  f"  인원: {record['최대_인원'] or '-'}명  구성: {record['구성'] or '-'}")

            results.append(record)

            if idx % 10 == 0:
                _save_progress(results)
                print(f"\n  💾 중간 저장 완료 ({idx}/{total})\n")

        await browser.close()

    if not results:
        print("수집된 데이터가 없습니다.")
        return
    print(f"\n[3단계] 엑셀 저장 중...")
    save_excel(results)

    print("\n" + "=" * 62)
    print(" 📊 Step 2 완료")
    print("=" * 62)
    print(f"  총 수집: {len(results)}개\n")

    priced = [r["1박_가격"] for r in results if isinstance(r.get("1박_가격"), int)]
    if priced:
        print(f"  평균 1박: ₩{sum(priced)//len(priced):,}")

    rates = [int(r["예약률_30일"].replace("%",""))
             for r in results if r.get("예약률_30일") and r["예약률_30일"] != ""]
    if rates:
        print(f"  평균 예약률: {sum(rates)//len(rates)}%")

    print(f"\n  ⚠️  미확인 필드 (값 없음) 비율 — {total}개 중:")
    for field, cnt in missing_counts.items():
        print(f"      {field}: {cnt}개 ({round(cnt/total*100)}%)")
    print(f"\n  → {OUTPUT_FILE.name}")
    print(f"\n  ※ 위 미확인 비율이 높다면 debug2_*.json 파일을 보고")
    print(f"     알려주세요 — 파서를 더 정확히 보강하겠습니다.")


def _save_progress(rows: list):
    pd.DataFrame(rows).to_csv(
        SAVE_DIR / "한옥_분석_진행중.csv",
        index=False, encoding="utf-8-sig"
    )


# ============================================================
# 📊  엑셀 저장
# ============================================================

def save_excel(rows: list):
    rows_sorted = sorted(
        rows,
        key=lambda x: (-(x.get("후기_수") or 0), -(float(x.get("별점") or 0)))
    )
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for i, r in enumerate(rows_sorted, 1):
        r["순위"]     = i
        r["수집_시각"] = now

    COLUMNS = [
        "순위", "구", "동", "숙소_이름", "링크",
        "호스트_이름", "호스트_응답률", "호스팅_시작",
        "별점", "후기_수", "슈퍼호스트", "게스트_픽",
        "최대_인원", "기준_인원", "구성", "최소_박수",
        "1박_가격", "주중_가격", "주말_가격", "청소비", "추가인원_요금",
        "예약률_30일",
        "에어컨", "주방", "세탁기", "주차", "와이파이", "TV", "욕조", "난방",
        "후기_키워드", "소개글", "수집_시각",
    ]
    COL_DISPLAY = {
        "순위": "순위", "구": "구(區)", "동": "동(洞)",
        "숙소_이름": "숙소 이름", "링크": "링크 (URL)",
        "호스트_이름": "호스트", "호스트_응답률": "응답률", "호스팅_시작": "호스팅 시작",
        "별점": "별점", "후기_수": "후기 수",
        "슈퍼호스트": "슈퍼호스트", "게스트_픽": "게스트 픽",
        "최대_인원": "최대 인원", "기준_인원": "기준 인원",
        "구성": "구성 (침실/침대/욕실)", "최소_박수": "최소 박수",
        "1박_가격": "1박 가격", "주중_가격": "주중 가격", "주말_가격": "주말 가격",
        "청소비": "청소비", "추가인원_요금": "추가인원 요금",
        "예약률_30일": "예약률 (30일)",
        "에어컨": "에어컨", "주방": "주방", "세탁기": "세탁기",
        "주차": "주차", "와이파이": "와이파이", "TV": "TV",
        "욕조": "욕조", "난방": "난방",
        "후기_키워드": "후기 키워드 Top5",
        "소개글": "숙소 소개글",
        "수집_시각": "수집 시각",
    }
    COL_WIDTHS = {
        "순위": 6, "구(區)": 9, "동(洞)": 12,
        "숙소 이름": 36, "링크 (URL)": 9,
        "호스트": 12, "응답률": 8, "호스팅 시작": 11,
        "별점": 7, "후기 수": 8, "슈퍼호스트": 9, "게스트 픽": 9,
        "최대 인원": 9, "기준 인원": 9,
        "구성 (침실/침대/욕실)": 20, "최소 박수": 9,
        "1박 가격": 12, "주중 가격": 12, "주말 가격": 12,
        "청소비": 11, "추가인원 요금": 13,
        "예약률 (30일)": 12,
        "에어컨": 7, "주방": 7, "세탁기": 7,
        "주차": 7, "와이파이": 8, "TV": 6, "욕조": 7, "난방": 7,
        "후기 키워드 Top5": 45,
        "숙소 소개글": 50,
        "수집 시각": 16,
    }

    df = pd.DataFrame(rows_sorted, columns=COLUMNS).rename(columns=COL_DISPLAY)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="한옥 분석")
        ws = writer.sheets["한옥 분석"]

        hfill  = PatternFill("solid", fgColor="2C3E50")
        hfont  = Font(bold=True, color="FFFFFF", size=10)
        thin   = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        amenity_cols = {"에어컨", "주방", "세탁기", "주차", "와이파이", "TV", "욕조", "난방"}

        for col_idx, col in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=col_idx)
            if col in amenity_cols:
                c.fill = PatternFill("solid", fgColor="1A6B3A")
            else:
                c.fill = hfill
            c.font = hfont
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin

        for row_idx in range(2, ws.max_row + 1):
            fill = (PatternFill("solid", fgColor="EBF5FB")
                    if row_idx % 2 == 0
                    else PatternFill("solid", fgColor="FFFFFF"))
            for col_idx in range(1, ws.max_column + 1):
                c   = ws.cell(row=row_idx, column=col_idx)
                hdr = ws.cell(row=1, column=col_idx).value
                c.fill, c.border = fill, thin
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if hdr == "링크 (URL)" and c.value:
                    c.hyperlink, c.value = c.value, "🔗"
                    c.font = Font(color="0563C1", underline="single")
                if c.value == "Y":
                    c.font = Font(bold=True, color="1A6B3A")

        for col_idx, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col, 8)
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "E2"

        summary_data = []
        for gu in df["구(區)"].unique():
            sub = df[df["구(區)"] == gu]
            prices = pd.to_numeric(sub["1박 가격"], errors="coerce").dropna()
            rates  = pd.to_numeric(
                sub["예약률 (30일)"].astype(str).str.replace("%", ""), errors="coerce"
            ).dropna()
            summary_data.append({
                "구": gu,
                "한옥 수": len(sub),
                "평균 별점": round(pd.to_numeric(sub["별점"], errors="coerce").mean(), 2),
                "평균 1박가격": round(prices.mean()) if len(prices) else "",
                "평균 예약률": f"{round(rates.mean())}%" if len(rates) else "",
                "슈퍼호스트 수": (sub["슈퍼호스트"] == "Y").sum(),
                "게스트픽 수": (sub["게스트 픽"] == "Y").sum(),
            })
        pd.DataFrame(summary_data).sort_values("한옥 수", ascending=False).to_excel(
            writer, index=False, sheet_name="구별 요약"
        )

    print(f"✅ 저장 완료: {OUTPUT_FILE.name}  ({len(rows)}개)")


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n\n⚠️  중단됨 — 진행 상황은 한옥_분석_진행중.csv에 저장되어 있습니다.")
