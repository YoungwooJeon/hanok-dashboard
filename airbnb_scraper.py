"""
============================================================
 에어비앤비 서울 한옥 스테이 전수 수집기 v4
============================================================
【방식】
  1. 브라우저로 에어비앤비 접속 → API 키 + 요청 본문 캡처
  2. 서울 5개 구(區)를 순회하며 API 직접 호출
  3. 한옥 키워드 포함 숙소만 필터링 + 중복 제거
  4. 각 숙소별 캘린더 API로 향후 30일 예약률 수집
  5. 엑셀 저장 (1박 가격 / 30일 예약률 포함)

【가격 계산】
  검색 시 내일~모레(2박) 날짜를 지정 → 총액 ÷ 2 = 1박 가격
  (최소 2박 요구 숙소 포함)

설치 (처음 한 번만):
    pip install playwright pandas openpyxl
    playwright install chromium
실행:
    python airbnb_scraper.py
============================================================
"""

import asyncio
import base64
import json
import re
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ============================================================
# ⚙️  설정
# ============================================================
CONFIG = {
    "max_pages_per_gu": 15,      # 구당 최대 페이지 (1페이지 = 18개)
    "headless":         False,
    "save_dir":         Path(__file__).parent,
    "delay_between_gu": 1.5,     # 구 전환 대기(초)
    "delay_calendar":   0.4,     # 캘린더 API 호출 간 대기(초)
}

# 수집 대상 구
SEOUL_GU = [
    "종로구",   # 북촌, 서촌, 익선동
    "중구",     # 명동, 을지로 인근
    "성북구",   # 성북동 고택
    "동대문구",
    "서대문구",
]

# 한옥 판별 키워드
HANOK_KEYWORDS = [
    "한옥", "hanok", "韓屋",
    "기와", "고택", "전통가옥", "전통 가옥",
    "한옥스테이", "한옥 스테이",
]

# 가격 계산용 날짜 (내일 ~ 모레, 2박)
_TODAY    = datetime.now()
CHECKIN   = (_TODAY + timedelta(days=1)).strftime("%Y-%m-%d")
CHECKOUT  = (_TODAY + timedelta(days=3)).strftime("%Y-%m-%d")
NIGHTS    = 2

# ============================================================
# 🌐  JavaScript: 검색 API 호출 (지역·날짜·커서 주입)
# ============================================================
JS_FETCH = """
async ([apiKey, requestBody, query, cursor, checkin, checkout]) => {
    try {
        const body   = JSON.parse(requestBody);
        const vars   = body.variables || {};
        const req    = vars.staysSearchRequest || vars.request || {};
        const params = req.rawParams || [];

        const setParam = (name, value) => {
            const found = params.find(p => p.filterName === name);
            if (found) { found.filterValues = [value]; }
            else        { params.push({ filterName: name, filterValues: [value] }); }
        };
        const removeParam = (name) => {
            const idx = params.findIndex(p => p.filterName === name);
            if (idx >= 0) params.splice(idx, 1);
        };

        setParam('query',    query);
        setParam('checkin',  checkin);
        setParam('checkout', checkout);
        setParam('adults',   '1');

        if (cursor) { setParam('cursor', cursor); }
        else        { removeParam('cursor'); }

        req.rawParams = params;

        const resp = await fetch(
            '/api/v3/StaysSearch?operationName=StaysSearch&locale=ko&currency=KRW',
            {
                method: 'POST',
                headers: {
                    'Content-Type':              'application/json',
                    'X-Airbnb-API-Key':          apiKey,
                    'X-Airbnb-GraphQL-Platform': 'web',
                    'Accept':                    'application/json',
                },
                body: JSON.stringify(body),
            }
        );
        if (!resp.ok) return { error: resp.status };
        return await resp.json();
    } catch(e) { return { error: e.toString() }; }
}
"""

# ============================================================
# 🗓️  JavaScript: 캘린더(예약 가능 여부) 조회
# ============================================================
JS_CALENDAR = """
async ([apiKey, listingId, year, month]) => {
    try {
        const url = `/api/v2/calendar_months?listing_id=${listingId}`
                  + `&month=${month}&year=${year}&count=2`
                  + `&_format=with_conditions&adults=1&children=0&infants=0&pets=0`;
        const resp = await fetch(url, {
            headers: {
                'X-Airbnb-API-Key': apiKey,
                'Accept':           'application/json',
            }
        });
        if (!resp.ok) return { error: resp.status };
        return await resp.json();
    } catch(e) { return { error: e.toString() }; }
}
"""

# ============================================================
# 🔍  응답 파싱
# ============================================================

def parse_results(data: dict) -> tuple[list[dict], list[str]]:
    listings, page_cursors = [], []
    try:
        results = data["data"]["presentation"]["staysSearch"]["results"]
    except (KeyError, TypeError):
        return listings, page_cursors

    page_cursors = (results.get("paginationInfo") or {}).get("pageCursors", [])

    for result in results.get("searchResults", []):
        # ID
        raw_id     = (result.get("demandStayListing") or {}).get("id", "")
        listing_id = ""
        if raw_id:
            try:    listing_id = base64.b64decode(raw_id).decode().split(":")[-1]
            except: listing_id = raw_id
        if not listing_id:
            continue

        # 이름
        name = (
            (result.get("nameLocalized") or {}).get("localizedStringWithTranslationPreference")
            or result.get("subtitle")
            or result.get("title") or ""
        )

        # 별점 / 후기
        rating, review_cnt = "", ""
        a11y = result.get("avgRatingA11yLabel", "")
        if a11y and a11y not in ("신규 숙소", "신규"):
            m = re.search(r"(\d+\.\d+)", a11y)
            if m: rating = m.group(1)
            m = re.search(r"후기\s*(\d[\d,]*)\s*개", a11y)
            if m: review_cnt = m.group(1).replace(",", "")

        # 1박 가격 (총액 ÷ NIGHTS)
        price_per_night = ""
        price_text = (
            (result.get("structuredDisplayPrice") or {})
            .get("primaryLine", {}).get("price", "")
        )
        if price_text:
            m = re.search(r"[\d,]+", price_text)
            if m:
                total = int(m.group(0).replace(",", ""))
                price_per_night = total // NIGHTS

        # 배지
        superhost, guest_fav = "", ""
        for badge in result.get("badges", []):
            t = badge.get("text", "")
            if "슈퍼호스트" in t or "Superhost" in t: superhost = "Y"
            if "게스트 선호" in t or "Guest favorite" in t: guest_fav = "Y"

        # 핵심 특징
        primary_lines = (result.get("structuredContent") or {}).get("primaryLine", [])
        features = " · ".join(
            x.get("body", "") for x in primary_lines if x.get("body")
        )

        listings.append({
            "listing_id":  listing_id,
            "숙소_이름":   name,
            "링크":        f"https://www.airbnb.co.kr/rooms/{listing_id}",
            "호스트_이름": "",
            "별점":        float(rating) if rating else "",
            "후기_수":     int(review_cnt) if review_cnt else "",
            "1박_가격":    price_per_night,
            "슈퍼호스트":  superhost,
            "게스트_픽":   guest_fav,
            "핵심_특징":   features,
            "예약률_30일": "",   # 캘린더 단계에서 채움
        })

    return listings, page_cursors


def is_hanok(listing: dict) -> bool:
    text = (listing.get("숙소_이름", "") + " " + listing.get("핵심_특징", "")).lower()
    return any(kw.lower() in text for kw in HANOK_KEYWORDS)


# ============================================================
# 🗓️  예약률 계산
# ============================================================

async def fetch_booking_rate(page, api_key: str, listing_id: str) -> str:
    """향후 30일 예약률 (%) 반환. 실패 시 빈 문자열."""
    today = _TODAY.date()
    result = await page.evaluate(
        JS_CALENDAR, [api_key, listing_id, today.year, today.month]
    )
    if not result or result.get("error"):
        return ""

    # 모든 날짜 수집
    all_days = []
    for cal_month in result.get("calendar_months", []):
        for day in cal_month.get("days", []):
            try:
                d     = datetime.strptime(day["date"], "%Y-%m-%d").date()
                avail = day.get("available", True)
                all_days.append((d, avail))
            except Exception:
                pass

    # 내일~30일 후
    start = today + timedelta(days=1)
    end   = today + timedelta(days=31)
    window = [(d, a) for d, a in all_days if start <= d < end]

    if len(window) < 20:   # 데이터 부족
        return ""

    booked = sum(1 for _, avail in window if not avail)
    return f"{round(booked / len(window) * 100)}%"


# ============================================================
# 📊  엑셀 저장
# ============================================================

COLUMNS = [
    "순위", "구", "숙소_이름", "링크", "호스트_이름",
    "별점", "후기_수", "1박_가격",
    "슈퍼호스트", "게스트_픽", "예약률_30일", "핵심_특징", "수집_시각",
]
COL_DISPLAY = {
    "순위":       "순위",
    "구":         "구(區)",
    "숙소_이름":  "숙소 이름",
    "링크":       "링크 (URL)",
    "호스트_이름":"호스트 이름",
    "별점":       "별점",
    "후기_수":    "후기 수",
    "1박_가격":   "1박 가격 (₩)",
    "슈퍼호스트": "슈퍼호스트",
    "게스트_픽":  "게스트 픽",
    "예약률_30일":"예약률 (30일)",
    "핵심_특징":  "핵심 특징",
    "수집_시각":  "수집 시각",
}
COL_WIDTHS = {
    "순위": 6, "구(區)": 10, "숙소 이름": 38, "링크 (URL)": 10,
    "호스트 이름": 14, "별점": 7, "후기 수": 8,
    "1박 가격 (₩)": 14, "슈퍼호스트": 10, "게스트 픽": 10,
    "예약률 (30일)": 12, "핵심 특징": 30, "수집 시각": 16,
}


def save_excel(rows: list[dict], path: Path):
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for i, r in enumerate(rows, 1):
        r["순위"]     = i
        r["수집_시각"] = now

    df = pd.DataFrame(rows, columns=COLUMNS).rename(columns=COL_DISPLAY)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="서울 한옥 스테이")
        ws = writer.sheets["서울 한옥 스테이"]

        hfill = PatternFill("solid", fgColor="C0392B")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        thin  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        for col_idx, col in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill, c.font = hfill, hfont
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin

        for row_idx in range(2, ws.max_row + 1):
            fill = (PatternFill("solid", fgColor="FDECEA")
                    if row_idx % 2 == 0
                    else PatternFill("solid", fgColor="FFFFFF"))
            for col_idx in range(1, ws.max_column + 1):
                c   = ws.cell(row=row_idx, column=col_idx)
                hdr = ws.cell(row=1, column=col_idx).value
                c.fill, c.border = fill, thin
                c.alignment = Alignment(vertical="center", wrap_text=True)
                if hdr == "링크 (URL)" and c.value:
                    c.hyperlink, c.value = c.value, "🔗 링크"
                    c.font = Font(color="0563C1", underline="single")

        for col_idx, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col, 14)

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # 구별 요약 시트
        summary = (
            df.groupby("구(區)").agg(
                숙소수       = ("숙소 이름", "count"),
                평균별점     = ("별점",     lambda x: round(pd.to_numeric(x, errors="coerce").mean(), 2)),
                평균1박가격  = ("1박 가격 (₩)", lambda x: round(pd.to_numeric(x, errors="coerce").mean())),
                평균예약률   = ("예약률 (30일)", lambda x: (
                    str(round(pd.to_numeric(x.str.replace("%",""), errors="coerce").mean())) + "%"
                    if x.str.replace("%","").apply(lambda v: v.isdigit()).any() else ""
                )),
                슈퍼호스트수 = ("슈퍼호스트", lambda x: (x == "Y").sum()),
                게스트픽수   = ("게스트 픽",   lambda x: (x == "Y").sum()),
            )
            .reset_index()
            .sort_values("숙소수", ascending=False)
        )
        summary.to_excel(writer, index=False, sheet_name="구별 요약")

    print(f"✅ 저장 완료 → {path.name}  ({len(rows)}개 한옥 숙소)")


# ============================================================
# 🚀  메인
# ============================================================

async def main():
    print("=" * 60)
    print(" 🏯 에어비앤비 서울 한옥 스테이 전수 수집기 v4")
    print(f"    대상: {', '.join(SEOUL_GU)}")
    print(f"    가격 기준: {CHECKIN} ~ {CHECKOUT} ({NIGHTS}박 ÷ {NIGHTS})")
    print("=" * 60)

    captured = {"api_key": None, "request_body": None}

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=CONFIG["headless"],
            args=["--disable-blink-features=AutomationControlled"],
        )
        ctx = await browser.new_context(
            viewport={"width": 1440, "height": 900},
            locale="ko-KR", timezone_id="Asia/Seoul",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        await ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
            "window.chrome={runtime:{}};"
        )
        page = await ctx.new_page()

        async def on_req(request):
            if "StaysSearch" in request.url and not captured["api_key"]:
                key = request.headers.get("x-airbnb-api-key", "")
                if key:
                    captured["api_key"] = key
                    try: captured["request_body"] = request.post_data
                    except: pass

        page.on("request", on_req)

        # ── 접속 ─────────────────────────────────────────────
        print("\n[1단계] 에어비앤비 접속 중...")
        await page.goto(
            "https://www.airbnb.co.kr/s/서울/homes?tab_id=home_tab&refinement_paths%5B%5D=%2Fhomes",
            wait_until="load", timeout=60000,
        )
        await asyncio.sleep(2)

        for _ in range(30):
            if captured["api_key"] and captured["request_body"]: break
            await asyncio.sleep(0.5)

        if not captured["api_key"]:
            print("❌ API 키 캡처 실패.")
            await browser.close()
            return
        print("        ✅ API 키 캡처 성공!")

        # ── 구별 수집 ─────────────────────────────────────────
        print(f"\n[2단계] {len(SEOUL_GU)}개 구 순회 + 한옥 필터링\n")

        all_hanok: dict[str, dict] = {}

        for gu_idx, gu in enumerate(SEOUL_GU, 1):
            query    = f"서울 {gu}"
            gu_total = 0
            gu_hanok = 0

            result = await page.evaluate(
                JS_FETCH,
                [captured["api_key"], captured["request_body"],
                 query, None, CHECKIN, CHECKOUT]
            )
            if not result or result.get("error"):
                print(f"  [{gu_idx}/{len(SEOUL_GU)}] {gu}: 오류 — 건너뜀")
                continue

            items, cursors = parse_results(result)
            gu_total += len(items)
            hanok_items = [x for x in items if is_hanok(x)]
            gu_hanok += len(hanok_items)
            for item in hanok_items:
                if item["listing_id"] not in all_hanok:
                    item["구"] = gu
                    all_hanok[item["listing_id"]] = item

            for p_idx in range(1, min(CONFIG["max_pages_per_gu"], len(cursors))):
                r2 = await page.evaluate(
                    JS_FETCH,
                    [captured["api_key"], captured["request_body"],
                     query, cursors[p_idx], CHECKIN, CHECKOUT]
                )
                if not r2 or r2.get("error"): break
                items2, _ = parse_results(r2)
                if not items2: break
                gu_total += len(items2)
                for item in items2:
                    if is_hanok(item):
                        gu_hanok += 1
                        if item["listing_id"] not in all_hanok:
                            item["구"] = gu
                            all_hanok[item["listing_id"]] = item
                await asyncio.sleep(0.5)

            print(f"  [{gu_idx}/{len(SEOUL_GU)}] {gu}: "
                  f"조회 {gu_total}개 → 한옥 {gu_hanok}개 (누적 {len(all_hanok)}개)")
            await asyncio.sleep(CONFIG["delay_between_gu"])

        if not all_hanok:
            print("\n❌ 한옥 숙소를 찾지 못했습니다.")
            await browser.close()
            return

        # ── 캘린더 (예약률) ───────────────────────────────────
        total = len(all_hanok)
        print(f"\n[3단계] {total}개 숙소 예약률 조회 중 (향후 30일)...\n")

        for i, (lid, item) in enumerate(all_hanok.items(), 1):
            rate = await fetch_booking_rate(page, captured["api_key"], lid)
            item["예약률_30일"] = rate
            print(f"\r  {i}/{total} 완료... ({item['숙소_이름'][:20]})", end="", flush=True)
            await asyncio.sleep(CONFIG["delay_calendar"])

        print(f"\n        ✅ 예약률 조회 완료!")

        await browser.close()

    # ── 저장 ─────────────────────────────────────────────────
    rows = list(all_hanok.values())
    rows.sort(key=lambda x: (-(x.get("후기_수") or 0), -(x.get("별점") or 0)))

    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = CONFIG["save_dir"] / f"airbnb_한옥_{ts}.xlsx"

    print(f"\n[4단계] 엑셀 저장 중...")
    save_excel(rows, out)

    # ── 요약 ─────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print(" 📊 최종 요약")
    print("=" * 60)
    print(f"  총 한옥 숙소:  {len(rows)}개")

    rated  = [float(r["별점"]) for r in rows if r.get("별점") != ""]
    if rated:
        print(f"  평균 별점:    {sum(rated)/len(rated):.2f}")

    priced = [r["1박_가격"] for r in rows if isinstance(r.get("1박_가격"), int)]
    if priced:
        print(f"  평균 1박:     ₩{sum(priced)//len(priced):,}")

    booked_rates = [int(r["예약률_30일"].replace("%",""))
                    for r in rows if r.get("예약률_30일") and r["예약률_30일"] != ""]
    if booked_rates:
        print(f"  평균 예약률:  {sum(booked_rates)//len(booked_rates)}%")

    print(f"\n  → {out.name}")


if __name__ == "__main__":
    asyncio.run(main())
